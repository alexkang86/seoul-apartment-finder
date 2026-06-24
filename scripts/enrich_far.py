#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""apartments.json 에 단지별 '용적률'(far)을 붙인다.
   출처: 건축HUB_건축물대장 표제부(getBrTitleInfo) — 용적률(vlRat)·건폐율(bcRat) 제공.
   K-apt 기본정보(data/..기본정보..xlsx)의 법정동주소에서 번지를 뽑고,
   data/bjd_code.csv(법정동명→10자리코드)로 sigunguCd/bjdongCd를 만들어 표제부를 조회한다.

   필요: data/.molit_key (건축물대장정보 서비스 활용신청 완료), openpyxl, data/bjd_code.csv
   환경변수: FAR_LIMIT(이번 실행 신규조회 상한, 기본 100000), FAR_DELAY(기본 0.1)
   캐시: data/far_cache.json {"sigungu|bjdong|bun|ji": 용적률or null}
"""
import os, csv, json, re, time, unicodedata, urllib.request, urllib.parse, urllib.error
from collections import defaultdict
def nfc(s): return unicodedata.normalize("NFC", s or "")

ROOT=os.path.join(os.path.dirname(__file__),"..")
APT=os.path.join(ROOT,"docs","data","apartments.json")
BJD=os.path.join(ROOT,"data","bjd_code.csv")
CACHE=os.path.join(ROOT,"data","far_cache.json")
URL="https://apis.data.go.kr/1613000/BldRgstHubService/getBrTitleInfo"
SEOUL_GU={"종로구","중구","용산구","성동구","광진구","동대문구","중랑구","성북구","강북구","도봉구","노원구","은평구","서대문구","마포구","양천구","강서구","구로구","금천구","영등포구","동작구","관악구","서초구","강남구","송파구","강동구"}
LIMIT=int(os.environ.get("FAR_LIMIT","100000")); DELAY=float(os.environ.get("FAR_DELAY","0.1"))

NAME_COLS=["단지명","kaptName"]; ADDR_COLS=["법정동주소","지번주소"]
SIDO_COLS=["시도"]; SGG_COLS=["시군구"]; DONG_COLS=["동리","읍면동","법정동"]

def load(p,d):
    try: return json.load(open(p,encoding="utf-8"))
    except Exception: return d
def col(row,cands):
    for c in cands:
        if c in row and row[c] not in (None,""): return str(row[c]).strip()
    return ""
def norm_name(s):
    s=re.sub(r"\(.*?\)","",s or ""); return re.sub(r"\s+","",s).replace("아파트","")

def read_basic_xlsx():
    import openpyxl, warnings; warnings.filterwarnings("ignore")
    # data/ 안에서 '기본정보' 엑셀 자동 선택(번지 있는 법정동주소 보유)
    dd=os.path.join(ROOT,"data")
    cand=[f for f in os.listdir(dd) if f.lower().endswith(".xlsx") and "기본정보" in nfc(f)]
    if not cand: cand=[f for f in os.listdir(dd) if f.lower().endswith(".xlsx") and "단지" in nfc(f)]
    if not cand: return []
    p=os.path.join(dd,cand[0])
    wb=openpyxl.load_workbook(p,data_only=True); ws=wb.active
    get=lambda ri:[ws.cell(row=ri,column=c).value for c in range(1,ws.max_column+1)]
    hr=1
    for ri in range(1,6):
        if any("단지명" in str(v or "") or "법정동" in str(v or "") for v in get(ri)): hr=ri; break
    hdr=[nfc(str(v or "")) for v in get(hr)]
    rows=[{hdr[i]:ws.cell(row=ri,column=i+1).value for i in range(len(hdr))} for ri in range(hr+1,ws.max_row+1)]
    wb.close()
    print(f"  기본정보: {os.path.basename(p)} {len(rows)}행")
    return rows

# 법정동주소에서 (대지구분, 번, 지) 추출.  예: "서울 종로구 내수동 72-3 …"→(0,0072,0003), "내수동 73-"→(0,0073,0000)
def parse_bunji(addr):
    if not addr: return None
    toks=addr.split()
    for i,t in enumerate(toks):
        if t.endswith(("동","가","리")) and i+1<len(toks):
            nxt=toks[i+1].replace("번지","")
            plat="0"
            if nxt=="산" and i+2<len(toks): plat="1"; nxt=toks[i+2]
            mm=re.match(r"^(\d+)(?:-(\d*))?$", nxt)   # 73- 처럼 끝에 '-'만 있어도 허용
            if mm: return (plat, f"{int(mm.group(1)):04d}", f"{int(mm.group(2) or 0):04d}")
    return None

# 경기 분구시 K-apt 표기('수원장안구')를 법정동코드 표기('수원시 장안구')로
GG_BORO=("수원","성남","안양","안산","고양","용인","부천")
def sgg_for_bjd(sido,sgg):
    if sido.startswith("경기") and "시" not in sgg and sgg.endswith("구"):
        for pre in GG_BORO:
            if sgg.startswith(pre): return pre+"시 "+sgg[len(pre):]
    return sgg

def fetch_far(key, sgg, bjd, plat, bun, ji):
    qs=urllib.parse.urlencode({"serviceKey":key,"sigunguCd":sgg,"bjdongCd":bjd,"platGbCd":plat,
        "bun":bun,"ji":ji,"numOfRows":100,"pageNo":1,"_type":"json"},safe="%")
    body=""
    for attempt in range(6):           # 서버가 빠른 호출엔 빈 200을 줌 → 백오프 재시도
        try:
            req=urllib.request.Request(URL+"?"+qs,headers={"User-Agent":"Mozilla/5.0","Accept":"application/json,*/*"})
            body=urllib.request.urlopen(req,timeout=25).read().decode("utf-8","ignore")
        except Exception:
            body=""
        if body.strip(): break
        time.sleep(0.8+0.4*attempt)
    if not body.strip(): return "EMPTY"   # 끝내 빈 응답 → 캐시 말고 다음 실행에 재시도
    try:
        j=json.loads(body)
        items=j.get("response",{}).get("body",{}).get("items","")
        if not items: return None
        it=items.get("item",[])
        if isinstance(it,dict): it=[it]
        # 같은 번지의 여러 동 중 '공동주택' 표제부만, 합리적 범위(30~1500%), 연면적 최대 동(주동)의 용적률
        best=None
        for x in it:
            if "공동주택" not in str(x.get("mainPurpsCdNm","")): continue
            try: vl=float(x.get("vlRat"))
            except: continue
            if not (30<=vl<=1500): continue
            try: area=float(x.get("totArea") or 0)
            except: area=0
            if not best or area>best[0]: best=(area, vl)
        return round(best[1]) if best else None
    except Exception:
        return None

def main():
    key=os.environ.get("BLDG_SERVICE_KEY")
    for kp in (os.path.join(ROOT,"data",".bldg_key"), os.path.join(ROOT,"data",".molit_key")):
        if not key and os.path.exists(kp): key="".join(open(kp).read().split())
    if not key: print("키 없음(data/.bldg_key)"); return
    if not os.path.exists(BJD): print("data/bjd_code.csv 없음"); return
    bjdmap={nfc(r["name"]):r["code"] for r in csv.DictReader(open(BJD,encoding="utf-8"))}
    cache=load(CACHE,{})
    rows=read_basic_xlsx()
    if not rows: print("기본정보 xlsx 없음 → data/ 에 K-apt 단지 기본정보 넣어주세요"); return

    # 단지 용적률 테이블: (sido,gu,dong)->{name_norm: far}
    table=defaultdict(dict); calls=0; resolved=0
    for r in rows:
        sido=nfc(col(r,SIDO_COLS)); sgg=nfc(col(r,SGG_COLS)); dong=nfc(col(r,DONG_COLS))
        nm=col(r,NAME_COLS); addr=nfc(col(r,ADDR_COLS))
        if not nm or not (sido.startswith("서울") or sido.startswith("경기")): continue
        full=" ".join(x for x in [sido,sgg_for_bjd(sido,sgg),dong] if x)
        code=bjdmap.get(full)
        if not code: continue
        pb=parse_bunji(addr)
        if not pb: continue
        plat,bun,ji=pb
        ckey=f"{code[:5]}|{code[5:]}|{bun}|{ji}|{plat}"
        if ckey in cache:
            far=cache[ckey]
        elif calls<LIMIT:
            far=fetch_far(key, code[:5], code[5:], plat, bun, ji)
            calls+=1
            if far=="EMPTY": far=None      # 빈 응답은 캐시 안 함(다음 실행 재시도)
            else: cache[ckey]=far
            if calls%200==0:
                print(f"  …조회 {calls}건"); json.dump(cache,open(CACHE,"w",encoding="utf-8"),ensure_ascii=False)
            time.sleep(DELAY)
        else:
            continue
        if far:
            # 우리 gu 표기로 환산: 서울=구, 경기=시+구
            if sido.startswith("서울"): gu=sgg
            else:
                parts=sgg.split(); si=next((t for t in parts if t.endswith("시")),"" ); gg=next((t for t in parts if t.endswith("구")),"")
                gu=(si[:-1] if si else "")+(gg[:-1] if gg else "")
            table[("서울" if sido.startswith("서울") else "경기", gu, dong)][norm_name(nm)]=far
            resolved+=1
    json.dump(cache,open(CACHE,"w",encoding="utf-8"),ensure_ascii=False)
    print(f"  표제부 조회 {calls}건(이번), 용적률 확보 단지 {resolved}건")

    # apartments.json 매칭
    d=load(APT,{"items":[]}); hit=0
    for a in d.get("items",[]): a.pop("far",None)   # 매번 새 스코어로 클린 재적용
    def apt_key(a):
        sido="서울" if a["gu"] in SEOUL_GU else "경기"
        dong=(a.get("dong") or "").split()[-1] if a.get("dong") else ""
        return (sido,a["gu"],dong)
    for a in d.get("items",[]):
        bucket=table.get(apt_key(a))
        if not bucket: continue
        nn=norm_name(a.get("name",""))
        far=bucket.get(nn)
        if far is None:
            cands=[v for k,v in bucket.items() if nn and (nn in k or k in nn)]
            far=max(cands) if cands else None
        if far: a["far"]=far; hit+=1
    json.dump(d,open(APT,"w",encoding="utf-8"),ensure_ascii=False,indent=1)
    print(f"용적률 매칭 완료: {hit}개 / 전체 {len(d.get('items',[]))}개 단지")

if __name__=="__main__": main()
